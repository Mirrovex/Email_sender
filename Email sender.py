import win32com.client as client
import os
import openpyxl
from tkinter import *
from tkinter import filedialog

pulpit = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

root = Tk()
root.title('Email sender')
root.geometry('650x800')
try:
    root.iconbitmap(".files\\email_send_ico.ico")
except:
    try:
        root.iconbitmap(os.path.join(sys.path[0], "email_send_ico.ico"))
    except:
        pass

#Okno dla wybierania folderu i pliku
filebrowser_frame = Frame(root)
filebrowser_frame.pack(pady = 30)

#wybieranie folderu
def dictbrowser():
    dictname = filedialog.askdirectory(initialdir = "/", title = "Wybierz folder z nazwiskami")
    if dictname != '':
        dict_name.set(dictname)
    else:
        dict_name.set('Nie wybrano folderu z nazwiskami')

dict_frame = Frame(filebrowser_frame)
dict_frame.grid(row = 0, column = 0)

dict_button = Button(dict_frame, text = 'Wybierz folder z nazwiskami', command = dictbrowser, cursor = 'hand2')
dict_button.grid(row = 0, column = 0, padx = 10)

dict_name = StringVar()
try:
    book = os.listdir(pulpit + '\\' + 'Kierowcy rozl czasu pracy')
    dict_name.set(pulpit + '\\' + 'Kierowcy rozl czasu pracy')
except:
    dict_name.set('Nie wybrano folderu z nazwiskami')

dict_label = Label(dict_frame, textvariable = dict_name)
dict_label.grid(row = 0, column = 1)

#Wybieranie pliku email
def filebrowser():
    filename = filedialog.askopenfilename(initialdir = "/", title = "Wybierz plik z email", filetype = (("Excel", ".xlsx"), ("Wszystkie pliki", ".*")))
    if filename != '':
        file_name.set(filename)
    else:
        file_name.set('Nie wybrano pliku z email')

file_frame = Frame(filebrowser_frame)
file_frame.grid(row = 1, column = 0, pady = 5)

file_button = Button(file_frame, text = 'Wybierz plik z email', command = filebrowser, cursor = 'hand2')
file_button.grid(row = 0, column = 0, padx = 10)

file_name = StringVar()
try:
    book = openpyxl.load_workbook(pulpit + '\\' + 'kierowcy email.xlsx')
    file_name.set(pulpit + '\\' + 'kierowcy email.xlsx')
except:
    file_name.set('Nie wybrano pliku z email')

file_label = Label(file_frame, textvariable = file_name)
file_label.grid(row = 0, column = 1)


#Temat
temat_frame = Frame(root)
temat_frame.pack()

temat_label = Label(temat_frame, text = "Temat: ")
temat_label.grid(row = 0, column = 0)

temat_txt = StringVar()
temat_txt.set('RCP miesięczne') #temat
temat_entry = Entry(temat_frame, textvariable = temat_txt, width = 83)
temat_entry.grid(row = 0, column = 1)

#Text
text_frame = Frame(root)
text_frame.pack(pady = 15)

scroll = Scrollbar(text_frame, cursor = 'hand2') #suwak
scroll.grid(row = 0, column = 2, sticky = NS)

text_text = Text(text_frame, width = 70, height = 27, wrap = WORD, yscrollcommand = scroll.set) #tresc
text_text.grid(row = 0, column = 0)
text_text.insert(INSERT, """Dzień dobry,

Wysyłam załączniki dot. „Rozliczenia czasu pracy kierowcy” za bieżący miesiąc:
-delegacje,
-ewidencje,
-raport ITD,
-raport PIP.

Proszę o informację zwrotną z potwierdzeniem otrzymania wyżej wymienionych dokumentów oraz ich akceptację o treści:
„Potwierdzam i akceptuję otrzymane dokumenty rozliczeniowe za bieżący miesiąc 2021”
- na e-mail:  rcp@lspgroup.pl

Proszę o odpisanie wciągu 7 dni od daty otrzymania.
Nie wywiązanie się z ustalonego terminu – skutkować będzie na wynik premii.


Pozdrawiam,
Karolina Nowicka
Specjalista ds. kadr

+48 784-694-941
+48 605-986-058
+48 22 123 68 51""") #tekst

scroll.config(command = text_text.yview) #aktualizuje pozycje suwaka

#Info
frame_info = Frame(root)
frame_info.pack(pady = 15)

info_txt1 = StringVar()
info_txt1.set('Aktualne informacje')
info_label1 = Label(frame_info, textvariable = info_txt1)
info_label1.grid(row = 0, column = 0)


#Główny kod
def send():
    if (dict_name.get() == 'Nie wybrano folderu z nazwiskami' or dict_name.get() == '') and (file_name.get() == 'Nie wybrano pliku z email' or file_name.get() == ''):
        info_txt1.set('Brak folderu głównego i pliku email')
        root.update()
    elif dict_name.get() == 'Nie wybrano folderu z nazwiskami' or dict_name.get() == '':
        info_txt1.set('Brak folderu z nazwiskami')
        root.update()
    elif file_name.get()  == 'Nie wybrano pliku z email' or file_name.get() == '':
        info_txt1.set('Brak pliku email')
        root.update()
    else: #jak wszystko jest dobrze
        book = openpyxl.load_workbook(file_name.get()) #bierze plik z wybranej ścieżki
        sheet = book.active
        x = 0 #ktory mail jest wysylany
        z = 1 #pozycja do zapisania bledu
        wszystkie_osoby = str(len(os.listdir(dict_name.get()))) #ilosc folderow w folderze glownym
        folder_glowny = dict_name.get()

        try:
            output = openpyxl.load_workbook(pulpit + '\\bledy email_sender.xlsx') #otworzenie pliku jesli istnieje
        except:
            output = openpyxl.Workbook() #jesli nie istnieje - stworzenie nowego
        sheet_output = output.active

        blad = 0 #ilosc niewyslanych maili

        info_txt1.set('Błędy')

        for osoba in os.listdir(folder_glowny): #bierze folder z podanej ścieżki

            info_txt1.set('Wysyłanie do: ' + osoba + ', wysyłanie: ' + str(x+1) + '/' + wszystkie_osoby)
            root.update()

            x += 1

            try:
                outlook = client.Dispatch("Outlook.Application")
                message = outlook.CreateItem(0)
                message.Display()
            except:
                info_txt1.set('Nie znaleziono aplikacji Outlook, lub jest zamknięta')
                root.update()
                x = 0 #ktory mail jest wysylany
                z = 1 #pozycja do zapisania bledu
                break

            email = ''

            for row in range(sheet.max_row): #szukanie emaila do nazwiska w folderze
                index = sheet[f'A{row + 1}'].value
                if index == None:
                    continue
                elif index.lower() == osoba.lower():
                    email = str(sheet[f'B{row + 1}'].value)
                    break

            if email == None or email == '': #zapisywanie braku emaila do excela
                blad += 1
                message.Close(1) #zamykanie strony outlooka
                sheet_output[f'A{z}'] = osoba
                sheet_output[f'B{z}'] = 'Brak maila'
                z += 1 #przejscie do nastepnego wiersza w bledach

            else:
                message.To = email
                message.Subject = temat_txt.get() #wstawianie tematu
                message.Body = text_text.get('1.0', END)

                for folder in os.listdir(folder_glowny + '\\' + osoba): #szukanie plikow

                        pdf = os.path.join(folder_glowny + '\\' + osoba + '\\' + folder)

                        message.Attachments.Add(pdf)

                message.Save()
                message.Send()
                
        else:
            if blad != 0: #jesli do kogos nie wyslano maila
                info_txt1.set('Nie wysłano do: ' + str(blad) + '/' + wszystkie_osoby + ' osób - brak maila')
                output.save(pulpit + '\\bledy email_sender.xlsx') #zapisanie excela z bledami
                output.close()
            else: #jesli do wszystkich wyslano maila
                info_txt1.set('Nie brakuje żadnego maila')
                output.close()
                os.remove(pulpit + '\\bledy email_sender.xlsx') #usuniecie pliku z bledami
            
        info_txt1.set(f'Wysłano do {x-blad}/{wszystkie_osoby} osób')
        book.close()




#Przycisk do wysylania
send_button = Button(root, text = 'Wyślij', font = ('Arial', 20), command = send, cursor = 'hand2')
send_button.pack(pady = 5)


root.mainloop()
